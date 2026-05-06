from __future__ import annotations

import math
import re
import warnings
from calendar import monthrange
from dataclasses import dataclass
from datetime import datetime
from typing import Optional

import pandas as pd
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Per-sheet header row overrides.
# Default is row index 5 (0-based), matching the original behaviour.
# Add overrides here if any sheet has its header on a different row.
# ---------------------------------------------------------------------------
SHEET_HEADER_ROW: dict[str, int] = {
    # sheet_name: 0-based row index of the header
    # e.g. '01_Advertiser_Name': 0,
}
DEFAULT_HEADER_ROW = 5


@dataclass
class DatabricksContext:
    path: str
    hash_name: str = ''
    tenant_id: str = ''
    account_id: str = ''
    window_start: object = None
    window_end: object = None
    downloaded: object = None
    window_days: Optional[int] = None
    ref_date: object = None
    ay: str = ''
    am: str = ''
    bn: str = ''
    al: str = ''
    au: str = ''
    bw: str = ''
    o7: object = None
    ax7: object = None
    journey_h7: object = None
    proj_h: object = None
    proj_i: object = None
    proj_j: object = None
    proj_k: object = None
    proj_cs_notes: str = ''
    df02: Optional[pd.DataFrame] = None
    df07: Optional[pd.DataFrame] = None
    df14: Optional[pd.DataFrame] = None
    df37: Optional[pd.DataFrame] = None
    df26: Optional[pd.DataFrame] = None
    df27: Optional[pd.DataFrame] = None
    df28: Optional[pd.DataFrame] = None
    df29: Optional[pd.DataFrame] = None
    df31: Optional[pd.DataFrame] = None
    df32: Optional[pd.DataFrame] = None
    df33: Optional[pd.DataFrame] = None
    df34: Optional[pd.DataFrame] = None
    df35: Optional[pd.DataFrame] = None
    metrics: dict = None
    parent_count: Optional[int] = None
    top1: Optional[float] = None
    top3: Optional[float] = None
    top5: Optional[float] = None
    tags: list = None
    gap: Optional[int] = None
    last_call: object = None
    prev_call: object = None


def clean_text(v) -> str:
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return ''
    return str(v).replace('&#39;', "'").strip()


def to_float(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)) and not pd.isna(v):
        return float(v)
    s = clean_text(v)
    if not s or s.lower() in {'nan', 'none', 'null', '-'}:
        return None
    s = s.replace('$', '').replace(',', '').strip()
    if s.endswith('%'):
        try:
            return float(s[:-1]) / 100.0
        except Exception:
            return None
    m = re.match(r'^([0-9]*\.?[0-9]+)k$', s, re.I)
    if m:
        return float(m.group(1)) * 1000.0
    try:
        return float(s)
    except Exception:
        return None


def norm_pct(v) -> Optional[float]:
    x = to_float(v)
    if x is None:
        return None
    return x if x <= 1 else x / 100.0


def pct_str(v: Optional[float], decimals: int = 1) -> str:
    if v is None:
        return 'Not documented'
    return f'{v * 100:.{decimals}f}%'


def money_str(v: Optional[float]) -> str:
    if v is None:
        return 'Not documented'
    return f'${v:,.0f}'


def trim(s: str, n: int = 260) -> str:
    s = re.sub(r'\s+', ' ', s or '').strip()
    return s if len(s) <= n else s[: n - 1].rstrip() + '…'


def _parse_header(ws) -> dict:
    """Extract account metadata from sheet 01_Advertiser_Name."""
    a1 = clean_text(ws['A1'].value)
    a2 = clean_text(ws['A2'].value)
    a3 = clean_text(ws['A3'].value)
    a4 = clean_text(ws['A4'].value)
    hash_name = re.sub(r'\s*-\s*Advertiser[_\s]*Name\s*$', '', a1, flags=re.I).strip()
    m2 = re.search(r'Tenant ID:\s*(.*?)\s*\|\s*Account ID:\s*(.*)$', a2)
    m3 = re.search(r'Date Range:\s*([0-9\-]+)\s*to\s*([0-9\-]+)', a3)
    tenant = m2.group(1).strip() if m2 else ''
    account_id = m2.group(2).strip() if m2 else ''
    start = datetime.strptime(m3.group(1), '%Y-%m-%d').date() if m3 else None
    end = datetime.strptime(m3.group(2), '%Y-%m-%d').date() if m3 else None
    downloaded = datetime.strptime(a4.replace('Downloaded:', '').strip(), '%Y-%m-%d %H:%M:%S') if a4 else None
    return {
        'hash_name': hash_name,
        'tenant_id': tenant,
        'account_id': account_id,
        'window_start': start,
        'window_end': end,
        'downloaded': downloaded,
        'window_days': (end - start).days + 1 if start and end else None,
        'ref_date': downloaded.date() if downloaded else end,
    }


def _find_sheet(wb, prefix: str) -> Optional[str]:
    """Return the first sheet name that starts with prefix, or None."""
    for name in wb.sheetnames:
        if name.startswith(prefix):
            return name
    return None


def _get_ws(wb, prefix: str):
    """Return the worksheet whose name starts with prefix, or None."""
    name = _find_sheet(wb, prefix)
    return wb[name] if name else None


def _get_df_from_wb(wb, sheet_prefix: str) -> Optional[pd.DataFrame]:
    """Read a sheet from an already-open workbook into a DataFrame using prefix matching."""
    sheet = _find_sheet(wb, sheet_prefix)
    if sheet is None:
        return None
    try:
        header_row = SHEET_HEADER_ROW.get(sheet_prefix, DEFAULT_HEADER_ROW)
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) <= header_row:
            return None
        headers = [str(c) if c is not None else f'Unnamed_{i}' for i, c in enumerate(rows[header_row])]
        data = rows[header_row + 1:]
        df = pd.DataFrame(data, columns=headers)
        return df
    except Exception:
        return None


def _ws_to_df(ws) -> Optional[pd.DataFrame]:
    """Convert an already-open worksheet to a DataFrame using DEFAULT_HEADER_ROW."""
    try:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) <= DEFAULT_HEADER_ROW:
            return None
        headers = [str(c) if c is not None else f'Unnamed_{i}' for i, c in enumerate(rows[DEFAULT_HEADER_ROW])]
        data = rows[DEFAULT_HEADER_ROW + 1:]
        df = pd.DataFrame(data, columns=headers)
        df = df.dropna(how='all')
        return df if not df.empty else None
    except Exception:
        return None


def _latest_row_by_modstamp(df: pd.DataFrame) -> pd.Series:
    """
    Return the row with the most recent SystemModstamp value.
    Falls back to the first row if the column is absent or unparseable.
    """
    modstamp_col = next(
        (c for c in df.columns if 'systemmod' in str(c).lower() or 'modstamp' in str(c).lower()),
        None
    )
    if modstamp_col:
        try:
            df = df.copy()
            df['_ts'] = pd.to_datetime(df[modstamp_col], errors='coerce')
            valid = df.dropna(subset=['_ts'])
            if not valid.empty:
                return valid.loc[valid['_ts'].idxmax()].drop(labels=['_ts'])
        except Exception:
            pass
    return df.iloc[0]


def _find_col_name(df: pd.DataFrame, *candidates: str) -> Optional[str]:
    """Return the first column name that matches any candidate (case-insensitive, ignoring _ and spaces)."""
    norm = {re.sub(r'[\s_]', '', str(c)).lower(): c for c in df.columns}
    for cand in candidates:
        key = re.sub(r'[\s_]', '', cand).lower()
        if key in norm:
            return norm[key]
    return None


def latest_gap_days(call_df: Optional[pd.DataFrame]):
    if call_df is None or call_df.empty or 'Gong__Call_End__c' not in call_df.columns:
        return None, None, None
    dates = pd.to_datetime(call_df['Gong__Call_End__c'], errors='coerce').dropna().sort_values()
    if len(dates) == 0:
        return None, None, None
    if len(dates) == 1:
        return None, dates.iloc[-1], None
    return int((dates.iloc[-1] - dates.iloc[-2]).days), dates.iloc[-1], dates.iloc[-2]


def load_databricks_context(path: str) -> DatabricksContext:
    # Open the workbook exactly once — all reads go through this handle.
    wb = load_workbook(path, data_only=True, read_only=True)

    try:
        ws01 = _get_ws(wb, '01_Advertiser_Name')
        if ws01 is None:
            raise ValueError('Sheet starting with 01_Advertiser_Name not found in export.')
        h = _parse_header(ws01)
        ctx = DatabricksContext(path=path, **h)

        # --- DataFrames ---
        ctx.df02 = _get_df_from_wb(wb, '02_Date_Range_KPIs__Date_Range_')
        ctx.df07 = _get_df_from_wb(wb, '07_KPIs_by_Parent_ASIN_by_Month')
        ctx.df14 = _get_df_from_wb(wb, '14_Campaign_Performance_by_Adve')
        ctx.df37 = _get_df_from_wb(wb, '37_Gong_Call_Insights_for_Sales')
        ctx.df26 = _get_df_from_wb(wb, '26_Unmanaged_ASIN')
        ctx.df27 = _get_df_from_wb(wb, '27_Timeframe_Boost')
        ctx.df28 = _get_df_from_wb(wb, '28_Unmanaged_Budget')
        ctx.df29 = _get_df_from_wb(wb, '29_Negative_Keywords__Global')
        ctx.df31 = _get_df_from_wb(wb, '31_Unmanaged_campaigns')
        ctx.df32 = _get_df_from_wb(wb, '32_Unmanaged_Campaigns_Budget_O')
        ctx.df33 = _get_df_from_wb(wb, '33_RBO_Configuration_Insights')
        ctx.df34 = _get_df_from_wb(wb, '34_Product_Level_ACoS')
        ctx.df35 = _get_df_from_wb(wb, '35_Campaign_Level_ACoS')

        # --- Tab 38: Client Success Insights — latest row by SystemModstamp ---
        ws38 = _get_ws(wb, '38_Client_Success_Insights_Repo')
        if ws38 is None:
            raise ValueError('Sheet starting with 38_Client_Success_Insights_Repo not found in export.')
        df38 = _ws_to_df(ws38)
        if df38 is None or df38.empty:
            import warnings
            warnings.warn(
                f'38_Client_Success_Insights_Repo has no data rows for {path}. '
                'Client success fields will be treated as missing.'
            )
            row38 = pd.Series([None] * 200)
        else:
            row38 = _latest_row_by_modstamp(df38)
        # Map by column position (headers may vary); fall back to positional index.
        # Original cells: AY7, AM7, BN7, AL7, AU7, BW7, O7, AX7
        # Column letters are 0-based index: A=0, O=14, AL=37, AM=38, AU=46,
        # AX=49, AY=50, BN=65, BW=75
        def _pos(letter: str):
            letter = letter.upper()
            idx = 0
            for ch in letter:
                idx = idx * 26 + (ord(ch) - ord('A') + 1)
            return idx - 1  # 0-based

        def _cell_val(row: pd.Series, letter: str):
            i = _pos(letter)
            return row.iloc[i] if i < len(row) else None

        ctx.ay  = clean_text(_cell_val(row38, 'AY'))
        ctx.am  = clean_text(_cell_val(row38, 'AM'))
        ctx.bn  = clean_text(_cell_val(row38, 'BN'))
        ctx.al  = clean_text(_cell_val(row38, 'AL'))
        ctx.au  = clean_text(_cell_val(row38, 'AU'))
        ctx.bw  = clean_text(_cell_val(row38, 'BW')).upper()
        ctx.o7  = _cell_val(row38, 'O')
        ctx.ax7 = _cell_val(row38, 'AX')

        # --- Tab 39 ---
        ws39 = _get_ws(wb, '39_Client_Journey_Insights_Data')
        ctx.journey_h7 = ws39['H7'].value if ws39 is not None else None

        # --- Tab 54: Project Dataset — filter by Advertiser_ID, latest by Modstamp ---
        ws54 = _get_ws(wb, '54_Project_Dataset_on_SF')
        if ws54 is None:
            raise ValueError('Sheet starting with 54_Project_Dataset_on_SF not found in export.')
        df54 = _ws_to_df(ws54)
        if df54 is None or df54.empty:
            import warnings
            warnings.warn(
                f'54_Project_Dataset_on_SF has no data rows for {path}. '
                'Project dataset fields will be treated as missing.'
            )
            row54 = pd.Series([None] * 200)
        else:
            # Filter to rows matching this export's Advertiser_ID
            adv_col = _find_col_name(df54, 'Advertiser_ID_c', 'Advertiser_ID', 'AdvertiserID')
            if adv_col and ctx.account_id:
                matched = df54[df54[adv_col].astype(str).str.strip() == str(ctx.account_id).strip()]
                df54_filtered = matched if not matched.empty else df54
            else:
                df54_filtered = df54

            # Pick latest row by SystemModstamp
            row54 = _latest_row_by_modstamp(df54_filtered)

        def _col54(letter: str):
            i = _pos(letter)
            return row54.iloc[i] if i < len(row54) else None

        ctx.proj_h        = _col54('H')
        ctx.proj_i        = _col54('I')
        ctx.proj_j        = _col54('J')
        ctx.proj_k        = _col54('K')
        ctx.proj_cs_notes = clean_text(_col54('T'))

    finally:
        try:
            wb.close()
        except Exception:
            pass

    # --- Derived metrics (no workbook needed) ---
    ctx.metrics = {}
    if ctx.df02 is not None and not ctx.df02.empty:
        row = ctx.df02.iloc[0]
        ctx.metrics = {k: row.get(k) for k in ['AdSales', 'TotalSales', 'AdSpend', 'TACoS', 'ACoS', 'Clicks', 'Revenue']}
        acos_val = row.get('ACoS')
        ctx.metrics['ROAS'] = (1 / float(acos_val)) if pd.notna(acos_val) and float(acos_val) != 0 else None

    if ctx.df07 is not None and not ctx.df07.empty and 'ParentASIN' in ctx.df07.columns and 'ThisYearTotalSales' in ctx.df07.columns:
        grp = ctx.df07.groupby('ParentASIN', dropna=True)['ThisYearTotalSales'].sum().sort_values(ascending=False)
        total = float(grp.sum()) if not grp.empty else 0.0
        ctx.parent_count = int(grp.index.nunique()) if not grp.empty else 0
        if total > 0:
            ctx.top1 = float(grp.head(1).sum() / total)
            ctx.top3 = float(grp.head(3).sum() / total)
            ctx.top5 = float(grp.head(5).sum() / total)

    ctx.tags = []
    if ctx.df14 is not None and not ctx.df14.empty:
        for c in ['Tag1', 'Tag2', 'Tag3', 'Tag4', 'Tag5']:
            if c in ctx.df14.columns:
                ctx.tags.extend([clean_text(x) for x in ctx.df14[c].dropna().tolist() if clean_text(x)])

    ctx.gap, ctx.last_call, ctx.prev_call = latest_gap_days(ctx.df37)
    return ctx


def monthly_budget_from_daily(ctx: DatabricksContext) -> Optional[float]:
    """
    Estimates monthly budget from the daily budget target in proj_h.
    Uses the month of window_end. Returns None with a warning if the
    window spans more than one calendar month, since the estimate would
    be misleading.
    """
    if ctx.window_end is None:
        return None
    daily = to_float(ctx.proj_h)
    if daily is None:
        return None
    if ctx.window_start is not None and ctx.window_start.month != ctx.window_end.month:
        warnings.warn(
            f"monthly_budget_from_daily: window spans "
            f"{ctx.window_start} to {ctx.window_end} (multiple months). "
            f"Budget estimate uses {ctx.window_end.strftime('%B %Y')} only and may be inaccurate.",
            stacklevel=2,
        )
    return daily * monthrange(ctx.window_end.year, ctx.window_end.month)[1]
