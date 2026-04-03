from __future__ import annotations
 
from openpyxl import load_workbook
from openpyxl.styles import Alignment
 
from reader_databricks_mastery import money_str, pct_str, clean_text
from rules_engine_mastery import interpretation
 
 
def write_mastery_output(template_path, output_path, summary, results, penalty, score, grade, findings, ctx):
    wb = load_workbook(template_path, keep_vba=True)
    ws_main = wb['Account Mastery_Analysis']
    ws_ref = wb['Account Mastery_Reference']
 
    # ── Page 1: Header block ───────────────────────────────────────────────────
    # Only write the dynamic account-level fields. Do not touch Key Findings
    # (rows 28+) — that section is fully driven by ArrayFormulas referencing
    # the Reference tab.
 
    ws_main['A1'] = f"{ctx.hash_name} — Account Mastery Analysis"
    ws_main['B3'] = f"Account: {ctx.hash_name} | Tenant ID: {ctx.tenant_id} | Account ID: {ctx.account_id}"
 
    if ctx.window_start and ctx.window_end and ctx.window_days:
        ws_main['B4'] = f"{ctx.window_start} to {ctx.window_end} ({ctx.window_days} days)"
 
    ws_main['B5'] = ctx.downloaded
    ws_main['B5'].number_format = 'yyyy-mm-dd hh:mm:ss'
 
    ws_main['C11'] = summary['primary_objective']
    ws_main['C13'] = summary.get('customization_context', '')
    ws_main['C16'] = money_str(summary['monthly_budget']) if summary['monthly_budget'] is not None else 'Monthly budget target not available.'
    ws_main['B18'] = pct_str(summary['acos_objective'])
    ws_main['B19'] = pct_str(summary['tacos_objective'])
    ws_main['B22'] = pct_str(summary['acos_constraint'])
    ws_main['B23'] = pct_str(summary['tacos_constraint'])
    ws_main['B24'] = money_str(summary['budget_constraint'])
    ws_main['E23'] = summary['primary_kpi']
    ws_main['D7'] = grade
    ws_main['D8'] = interpretation(grade)
 
    for cell in ['C11', 'C13', 'C16', 'D8']:
        ws_main[cell].alignment = Alignment(wrap_text=True, vertical='top')
 
    # ── Page 2: Reference tab — STATUS, What We Saw, Why It Matters only ──────
    # The template already contains: Control Name (C), What/Why/How definitions
    # (E/F/G), Data Source (J), Importance (L), and formula-driven Priority (M).
    # The agent writes only columns D, H, I — nothing else.
 
    cid_to_row = {}
    for r in range(2, ws_ref.max_row + 1):
        cid = clean_text(ws_ref[f'B{r}'].value).upper()
        if cid:
            cid_to_row[cid] = r
 
    for cid, res in results.items():
        if cid not in cid_to_row:
            print(f"[writer] WARNING: control {cid} not found in Account Mastery_Reference sheet — skipping.")
            continue
        rr = cid_to_row[cid]
 
        ws_ref[f'D{rr}'] = res.status          # STATUS
        ws_ref[f'H{rr}'] = res.what            # What We Saw
        ws_ref[f'I{rr}'] = res.why             # Why It Matters
 
        for cell_ref in [f'H{rr}', f'I{rr}']:
            ws_ref[cell_ref].alignment = Alignment(wrap_text=True, vertical='top')
 
    # ── Logic and Calculation: no writes ──────────────────────────────────────
    # All scoring (penalty, score, grade) is formula-driven via references to
    # the Reference tab. Writing Python-computed values here would overwrite
    # those formulas and break recalculation.
 
    wb.save(output_path)
    try:
        wb.close()
    except Exception:
        pass
